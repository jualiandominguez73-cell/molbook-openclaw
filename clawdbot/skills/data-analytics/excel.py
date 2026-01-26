#!/usr/bin/env python3
"""
Data Analytics - Excel Processing Script
Read/write Excel files, list sheets, and perform Excel operations.
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def cmd_read(args):
    """Read Excel file and display contents."""
    try:
        # Read specific sheet or default
        if args.sheet:
            df = pd.read_excel(args.file, sheet_name=args.sheet)
        else:
            df = pd.read_excel(args.file)
        
        print(f"=== {args.file} ===")
        if args.sheet:
            print(f"Sheet: {args.sheet}")
        print(f"Shape: {len(df)} rows × {len(df.columns)} columns")
        print(f"\nColumns: {', '.join(df.columns)}")
        
        if args.head:
            print(f"\nFirst {args.head} rows:")
            print(df.head(args.head).to_string())
        else:
            print(f"\nData:")
            print(df.to_string())
        
        if args.output:
            df.to_csv(args.output, index=False)
            print(f"\nExported to {args.output}")
            
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)


def cmd_write(args):
    """Write data to Excel file."""
    try:
        # Parse data from JSON
        if args.data:
            data = json.loads(args.data)
            df = pd.DataFrame(data)
        elif args.csv:
            df = pd.read_csv(args.csv)
        else:
            print("Error: Provide --data (JSON) or --csv (file path)")
            sys.exit(1)
        
        # Write to Excel
        sheet_name = args.sheet if args.sheet else 'Sheet1'
        df.to_excel(args.file, sheet_name=sheet_name, index=False)
        
        print(f"Written {len(df)} rows to {args.file} (sheet: {sheet_name})")
        
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON data: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error writing Excel file: {e}")
        sys.exit(1)


def cmd_list_sheets(args):
    """List all sheets in an Excel workbook."""
    try:
        wb = load_workbook(args.file, read_only=True)
        sheets = wb.sheetnames
        
        print(f"=== Sheets in {args.file} ===")
        for i, sheet in enumerate(sheets, 1):
            # Get row count
            ws = wb[sheet]
            row_count = ws.max_row if ws.max_row else 0
            col_count = ws.max_column if ws.max_column else 0
            print(f"{i}. {sheet} ({row_count} rows × {col_count} cols)")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)


def cmd_merge(args):
    """Merge/join two Excel files or sheets."""
    try:
        # Read both files
        df1 = pd.read_excel(args.file1, sheet_name=args.sheet1 if args.sheet1 else 0)
        df2 = pd.read_excel(args.file2, sheet_name=args.sheet2 if args.sheet2 else 0)
        
        # Perform merge (VLOOKUP equivalent)
        how = args.how if args.how else 'left'
        result = pd.merge(df1, df2, on=args.on, how=how)
        
        print(f"=== Merge Result ===")
        print(f"Left: {len(df1)} rows, Right: {len(df2)} rows")
        print(f"Result: {len(result)} rows ({how} join on '{args.on}')")
        print(result.to_string())
        
        if args.output:
            result.to_excel(args.output, index=False)
            print(f"\nSaved to {args.output}")
            
    except Exception as e:
        print(f"Error in merge: {e}")
        sys.exit(1)


def cmd_info(args):
    """Show detailed info about Excel file."""
    try:
        wb = load_workbook(args.file, read_only=True)
        
        print(f"=== Excel File Info ===")
        print(f"File: {args.file}")
        print(f"Sheets: {len(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- {sheet_name} ---")
            print(f"  Dimensions: {ws.dimensions}")
            print(f"  Max row: {ws.max_row}")
            print(f"  Max col: {ws.max_column}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description='Data Analytics - Excel Tool',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    subparsers = parser.add_subparsers(dest='command', help='Commands')
    
    # read command
    read_parser = subparsers.add_parser('read', help='Read Excel file')
    read_parser.add_argument('file', help='Excel file path')
    read_parser.add_argument('--sheet', '-s', help='Sheet name to read')
    read_parser.add_argument('--head', type=int, help='Show only first N rows')
    read_parser.add_argument('--output', '-o', help='Export to CSV')
    
    # write command
    write_parser = subparsers.add_parser('write', help='Write to Excel file')
    write_parser.add_argument('file', help='Output Excel file path')
    write_parser.add_argument('--data', '-d', help='Data as JSON (list of dicts or dict of lists)')
    write_parser.add_argument('--csv', help='Read data from CSV file')
    write_parser.add_argument('--sheet', '-s', help='Sheet name (default: Sheet1)')
    
    # list-sheets command
    list_parser = subparsers.add_parser('list-sheets', help='List sheets in workbook')
    list_parser.add_argument('file', help='Excel file path')
    
    # merge command (VLOOKUP equivalent)
    merge_parser = subparsers.add_parser('merge', help='Merge two Excel files (VLOOKUP)')
    merge_parser.add_argument('file1', help='First Excel file')
    merge_parser.add_argument('file2', help='Second Excel file')
    merge_parser.add_argument('--on', required=True, help='Column to join on')
    merge_parser.add_argument('--sheet1', help='Sheet from file1')
    merge_parser.add_argument('--sheet2', help='Sheet from file2')
    merge_parser.add_argument('--how', choices=['left', 'right', 'inner', 'outer'], default='left')
    merge_parser.add_argument('--output', '-o', help='Output file')
    
    # info command
    info_parser = subparsers.add_parser('info', help='Show Excel file info')
    info_parser.add_argument('file', help='Excel file path')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    commands = {
        'read': cmd_read,
        'write': cmd_write,
        'list-sheets': cmd_list_sheets,
        'merge': cmd_merge,
        'info': cmd_info,
    }
    
    commands[args.command](args)


if __name__ == '__main__':
    main()
